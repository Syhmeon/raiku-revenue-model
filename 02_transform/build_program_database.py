"""
Build Program Database — Merge Dune + SolWatch per-program data
================================================================
Merges per-program analytics from multiple sources into a clean CSV.

Sources:
  1. Dune aggregate (30-day): fee/CU, priority fees, CU — primary
  2. Dune 7-day snapshot: fee/CU by program — fallback if 30-day not yet available
  3. SolWatch / lead_pipeline: tx count, fail rate, names, categories — enrichment

Output: data/processed/program_database.csv (semicolon-delimited)

The program_categories.csv mapping (data/mapping/) is applied to classify
programs into RAIKU archetypes.
"""

import csv
import sys
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DATA_RAW, DATA_PROCESSED, DATA_MAPPING, CSV_DELIMITER, CSV_ENCODING

# ── Input files ────────────────────────────────────────────
DUNE_AGGREGATE_FILE = DATA_RAW / "dune_program_fees_aggregate.csv"
DUNE_7DAY_FILE = DATA_RAW / "dune_fee_per_cu_by_program.csv"
SOLWATCH_FILE = DATA_RAW / "lead_pipeline_sheet.xlsx"
MAPPING_FILE = DATA_MAPPING / "program_categories.csv"

# ── Output ─────────────────────────────────────────────────
OUTPUT_FILE = DATA_PROCESSED / "program_database.csv"

OUTPUT_COLUMNS = [
    "program_id",               # A - Address (PK)
    "program_name",             # B - Human-readable name
    "raiku_category",           # C - RAIKU archetype category
    "raiku_subcategory",        # D - Subcategory (amm, orderbook, prop_amm, etc.)
    "raiku_product",            # E - aot / jit / both / potential / neither / unknown
    "data_source",              # F - Which source provided the data
    "period_days",              # G - Data period in days
    "tx_count",                 # H - Total transactions
    "success_count",            # I - Successful transactions
    "fail_rate",                # J - Failure rate (0-1)
    "total_fees_sol",           # K - Total fees (SOL)
    "priority_fees_sol",        # L - Priority fees (SOL)
    "total_cu",                 # M - Total CU consumed
    "avg_cu_per_tx",            # N - Average CU per transaction
    "median_fee_per_cu",        # O - Median fee/CU (lamports)
    "p25_fee_per_cu",           # P - P25 fee/CU (lamports)
    "p75_fee_per_cu",           # Q - P75 fee/CU (lamports)
    "avg_fee_per_cu_lamports",  # R - Average fee/CU (lamports)
    "pct_of_total_priority",    # S - % of total priority fees (computed)
    "solwatch_pain_score",      # T - SolWatch pain score (if available)
    "solwatch_fail_rate",       # U - SolWatch fail rate (if available)
]


def safe_float(val, default=None):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def load_csv(filepath):
    if not filepath.exists():
        print(f"  WARNING: {filepath.name} not found")
        return []
    with open(filepath, "r", encoding=CSV_ENCODING) as f:
        content = f.read()
        if content.startswith("\ufeff"):
            content = content[1:]
        reader = csv.DictReader(content.splitlines(), delimiter=CSV_DELIMITER)
        return list(reader)


def load_mapping():
    """Load program → category mapping."""
    if not MAPPING_FILE.exists():
        print(f"  WARNING: Mapping file not found ({MAPPING_FILE.name})")
        return {}
    rows = load_csv(MAPPING_FILE)
    mapping = {}
    for r in rows:
        pid = r.get("program_id", "").strip()
        if pid:
            mapping[pid] = {
                "name": r.get("program_name", ""),
                "category": r.get("raiku_category", "unknown"),
                "subcategory": r.get("subcategory", ""),
                "product": r.get("raiku_product", "unknown"),
            }
    print(f"    Mapping: {len(mapping)} programs classified")
    return mapping


def load_solwatch():
    """Load SolWatch data from lead_pipeline Excel.
    Returns dict: program_id → {name, category, pain_score, fail_rate, ...}
    """
    if not SOLWATCH_FILE.exists():
        print(f"  WARNING: SolWatch file not found ({SOLWATCH_FILE.name})")
        return {}

    try:
        import openpyxl
    except ImportError:
        print("  WARNING: openpyxl not installed, skipping SolWatch data")
        print("  Install with: pip install openpyxl")
        return {}

    wb = openpyxl.load_workbook(SOLWATCH_FILE, read_only=True, data_only=True)

    # Load Programs tab for names and categories
    programs = {}
    if "Programs" in wb.sheetnames:
        ws = wb["Programs"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            for row in rows[1:]:
                data = dict(zip(headers, row))
                pid = str(data.get("program id", data.get("program_id", ""))).strip()
                if pid:
                    programs[pid] = {
                        "name": str(data.get("name", data.get("program name", ""))).strip(),
                        "category": str(data.get("category", "unknown")).strip(),
                    }

    # Load Data tab for aggregated metrics
    solwatch = {}
    if "Data" in wb.sheetnames:
        ws = wb["Data"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            for row in rows[1:]:
                data = dict(zip(headers, row))
                pid = str(data.get("program id", data.get("program_id", ""))).strip()
                if not pid:
                    continue
                if pid not in solwatch:
                    solwatch[pid] = {
                        "name": programs.get(pid, {}).get("name", ""),
                        "category": programs.get(pid, {}).get("category", "unknown"),
                        "pain_score": safe_float(data.get("pain score")),
                        "fail_rate": safe_float(data.get("fail rate")),
                        "tx_count": 0,
                        "priority_fees_sol": 0,
                    }
                solwatch[pid]["tx_count"] += safe_float(data.get("tx count"), 0)
                solwatch[pid]["priority_fees_sol"] += safe_float(data.get("priority fees (sol)"), 0)

    wb.close()
    print(f"    SolWatch: {len(solwatch)} programs loaded")
    return solwatch


def build():
    print("  Loading sources...")

    # 1. Load Dune aggregate (preferred) or 7-day fallback
    dune_rows = load_csv(DUNE_AGGREGATE_FILE)
    source_label = "dune_30d"
    period = 30

    if not dune_rows:
        print("  → Falling back to 7-day Dune data")
        dune_rows = load_csv(DUNE_7DAY_FILE)
        source_label = "dune_7d"
        period = 7

    if dune_rows:
        print(f"    Dune ({source_label}): {len(dune_rows)} programs")
    else:
        print("    Dune: no data available")

    # 2. Load mapping
    mapping = load_mapping()

    # 3. Load SolWatch
    solwatch = load_solwatch()

    # 4. Build merged database
    print("\n  Merging...")
    programs = {}

    # Dune is primary source
    for r in dune_rows:
        pid = r.get("program_id", r.get("primary_program", "")).strip()
        if not pid:
            continue

        # Name: from mapping > SolWatch > Dune
        name = (mapping.get(pid, {}).get("name")
                or solwatch.get(pid, {}).get("name")
                or r.get("program_name", ""))

        # Category: from mapping > SolWatch
        cat = (mapping.get(pid, {}).get("category")
               or solwatch.get(pid, {}).get("category", "unknown"))
        subcat = mapping.get(pid, {}).get("subcategory", "")
        product = mapping.get(pid, {}).get("product", "unknown")

        tx = safe_float(r.get("tx_count"), 0)
        success = safe_float(r.get("success_count"), 0)

        programs[pid] = {
            "program_id": pid,
            "program_name": name,
            "raiku_category": cat,
            "raiku_subcategory": subcat,
            "raiku_product": product,
            "data_source": source_label,
            "period_days": period,
            "tx_count": int(tx) if tx else "",
            "success_count": int(success) if success else "",
            "fail_rate": round(1 - success / tx, 4) if tx > 0 else "",
            "total_fees_sol": safe_float(r.get("total_fees_sol")),
            "priority_fees_sol": safe_float(r.get("priority_fees_sol")),
            "total_cu": safe_float(r.get("total_cu")),
            "avg_cu_per_tx": safe_float(r.get("avg_cu_consumed")),
            "median_fee_per_cu": safe_float(r.get("median_fee_per_cu")),
            "p25_fee_per_cu": safe_float(r.get("p25_fee_per_cu")),
            "p75_fee_per_cu": safe_float(r.get("p75_fee_per_cu")),
            "avg_fee_per_cu_lamports": safe_float(r.get("avg_fee_per_cu_lamports")),
            "pct_of_total_priority": None,  # Computed below
            "solwatch_pain_score": solwatch.get(pid, {}).get("pain_score"),
            "solwatch_fail_rate": solwatch.get(pid, {}).get("fail_rate"),
        }

    # Compute % of total priority fees
    total_pf = sum(safe_float(p.get("priority_fees_sol"), 0) for p in programs.values())
    if total_pf > 0:
        for p in programs.values():
            pf = safe_float(p.get("priority_fees_sol"), 0)
            p["pct_of_total_priority"] = round(pf / total_pf, 6) if pf > 0 else 0

    # Sort by priority fees descending
    sorted_programs = sorted(programs.values(),
                             key=lambda p: safe_float(p.get("priority_fees_sol"), 0),
                             reverse=True)

    # Save
    DATA_PROCESSED.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding=CSV_ENCODING, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, delimiter=CSV_DELIMITER,
                                extrasaction="ignore")
        writer.writeheader()
        for p in sorted_programs:
            clean = {k: ("" if v is None else v) for k, v in p.items()}
            writer.writerow(clean)

    print(f"\n  Saved: {OUTPUT_FILE}")
    print(f"  {len(sorted_programs)} programs × {len(OUTPUT_COLUMNS)} columns")

    # Stats
    classified = sum(1 for p in sorted_programs if p.get("raiku_category") not in ("unknown", ""))
    print(f"\n  Classification: {classified}/{len(sorted_programs)} programs mapped")
    if total_pf > 0:
        top10_pf = sum(safe_float(p.get("priority_fees_sol"), 0) for p in sorted_programs[:10])
        print(f"  Top 10 = {top10_pf/total_pf*100:.1f}% of total priority fees")

    # Show top 10
    print(f"\n  Top 10 programs by priority fees ({source_label}):")
    for i, p in enumerate(sorted_programs[:10], 1):
        name = p.get("program_name") or p["program_id"][:12] + "..."
        pf = safe_float(p.get("priority_fees_sol"), 0)
        cat = p.get("raiku_category", "?")
        med = safe_float(p.get("median_fee_per_cu"))
        med_str = f"{med:.4f}" if med is not None else "?"
        print(f"    {i:>2}. {name:<30} | {pf:>10.2f} SOL | med fee/CU: {med_str} | cat: {cat}")


if __name__ == "__main__":
    print("\n=== Building Program Database ===")
    build()
    print("\nDone.")
